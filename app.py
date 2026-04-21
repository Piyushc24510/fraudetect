from flask import Flask, render_template, request, session, redirect, url_for, flash, jsonify
from functools import wraps
import pickle
import os
from datetime import datetime, timedelta
import hashlib
import re
import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.sentiment import SentimentIntensityAnalyzer
import json
from collections import Counter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import sqlite3

# FLASK APP SETUP
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "fraud-detection-dev-key-change-in-production")

# DATABASE SETUP
DATABASE = os.environ.get(
    "DATABASE_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "fraud_detection.db")
)
def get_db():
    """Get database connection"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn
def init_db():
    """Initialize database with tables"""
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                name TEXT NOT NULL,
                email TEXT,
                role TEXT DEFAULT 'user',
                joined TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                review TEXT NOT NULL,
                review_full TEXT,
                is_fake INTEGER NOT NULL,
                confidence REAL NOT NULL,
                checked_by TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        # Create default admin if not exists
        cursor.execute("SELECT id FROM users WHERE username = 'admin'")
        if not cursor.fetchone():
            admin_pass = hashlib.sha256("admin123".encode()).hexdigest()
            cursor.execute(
                "INSERT INTO users (username, password, name, role, joined) VALUES (?, ?, ?, ?, ?)",
                ("admin", admin_pass, "Administrator", "admin", datetime.now().strftime("%Y-%m-%d"))
            )
        conn.commit()
    finally:
        conn.close()
init_db()
# EMAIL NOTIFICATION CONFIG
NOTIFY_EMAIL_SENDER   = os.environ.get("NOTIFY_EMAIL_SENDER", "")
NOTIFY_EMAIL_PASSWORD = os.environ.get("NOTIFY_EMAIL_PASSWORD", "")
NOTIFY_EMAIL_RECEIVER = os.environ.get("NOTIFY_EMAIL_RECEIVER", "")
NOTIFY_EMAIL_ENABLED  = os.environ.get("NOTIFY_EMAIL_ENABLED", "false").lower() == "true"
# HELPER FUNCTIONS
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated_function
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        if session.get("user_role") != "admin":
            flash("Admin access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated_function
def get_user_history(username=None):
    """Get history from database"""
    conn = get_db()
    try:
        cursor = conn.cursor()
        if username:
            cursor.execute("SELECT * FROM history WHERE checked_by = ? ORDER BY id DESC", (username,))
        else:
            cursor.execute("SELECT * FROM history ORDER BY id DESC")
        return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()
def add_to_history(review, review_full, is_fake, confidence, checked_by):
    """Add review to database history"""
    conn = get_db()
    try:
        cursor = conn.cursor()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "INSERT INTO history (review, review_full, is_fake, confidence, checked_by, timestamp) VALUES (?, ?, ?, ?, ?, ?)",
            (review, review_full, is_fake, confidence, checked_by, timestamp)
        )
        conn.commit()
    finally:
        conn.close()
# EMAIL NOTIFICATION FUNCTION
def send_registration_email(username, name, role):
    """Send email notification when a new user registers"""
    if not NOTIFY_EMAIL_ENABLED:
        return
    if not all([NOTIFY_EMAIL_SENDER, NOTIFY_EMAIL_PASSWORD, NOTIFY_EMAIL_RECEIVER]):
        print("⚠️ Email not sent: missing sender/password/receiver config")
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🔔 New User Registration — {name} ({username})"
        msg["From"]    = NOTIFY_EMAIL_SENDER
        msg["To"]      = NOTIFY_EMAIL_RECEIVER
        registered_at = datetime.now().strftime("%d %b %Y, %I:%M %p")
        html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0a0a0a;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#0a0a0a;padding:40px 20px;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0"
             style="background:linear-gradient(135deg,#1e1e1e,#2a2a2a);
                    border-radius:16px;border:1px solid #333;
                    box-shadow:0 0 30px rgba(0,255,153,0.15);max-width:600px;width:100%;">
        <tr>
          <td style="background:linear-gradient(90deg,#00ff99,#00cc7a);
                     padding:28px 36px;text-align:center;">
            <h1 style="margin:0;color:#000;font-size:22px;font-weight:700;">
              🛡️ Review Fraud Detection System
            </h1>
            <p style="margin:6px 0 0;color:#003322;font-size:13px;">New User Registration Alert</p>
          </td>
        </tr>
        <tr>
          <td style="padding:36px;">
            <p style="margin:0 0 24px;color:#aaa;font-size:14px;">A new account has been created:</p>
            <table width="100%" style="background:#111;border-radius:12px;border:1px solid #333;margin-bottom:28px;">
              <tr style="border-bottom:1px solid #222;">
                <td style="padding:14px 20px;color:#666;font-size:13px;width:38%;">👤 Full Name</td>
                <td style="padding:14px 20px;color:#fff;font-size:14px;font-weight:600;">{name}</td>
              </tr>
              <tr style="border-bottom:1px solid #222;">
                <td style="padding:14px 20px;color:#666;font-size:13px;">🔑 Username</td>
                <td style="padding:14px 20px;color:#00ff99;font-size:14px;font-weight:600;
                           font-family:monospace;">{username}</td>
              </tr>
              <tr style="border-bottom:1px solid #222;">
                <td style="padding:14px 20px;color:#666;font-size:13px;">🎭 Role</td>
                <td style="padding:14px 20px;">
                  <span style="background:#1a3a2a;color:#00ff99;padding:4px 12px;
                               border-radius:20px;font-size:12px;font-weight:700;
                               text-transform:uppercase;">{role}</span>
                </td>
              </tr>
              <tr>
                <td style="padding:14px 20px;color:#666;font-size:13px;">🕐 Registered At</td>
                <td style="padding:14px 20px;color:#fff;font-size:13px;">{registered_at}</td>
              </tr>
            </table>
          </td>
        </tr>
        <tr>
          <td style="background:#111;padding:20px 36px;border-top:1px solid #222;text-align:center;">
            <p style="margin:0;color:#444;font-size:12px;">Automated notification from Review Fraud Detection System</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>
"""
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(NOTIFY_EMAIL_SENDER, NOTIFY_EMAIL_PASSWORD)
            server.sendmail(NOTIFY_EMAIL_SENDER, NOTIFY_EMAIL_RECEIVER, msg.as_string())
        print(f"✅ Registration email sent for user: {username}")
    except Exception as e:
        print(f"❌ Email error (non-critical): {e}")
# LOAD ML MODEL
# ==========================================
def load_model():
    """Load model from multiple possible locations"""
    search_paths = [
        ("model.pkl", "vectorizer.pkl"),
        ("/tmp/model.pkl", "/tmp/vectorizer.pkl"),
    ]
    for model_path, vec_path in search_paths:
        try:
            m = pickle.load(open(model_path, "rb"))
            v = pickle.load(open(vec_path, "rb"))
            print(f"✅ ML Model loaded from {model_path}")
            return m, v
        except Exception:
            continue
    print("⚠️ Model not found - run train_model.py first")
    return None, None
model, vectorizer = load_model()
# NLTK SETUP
nltk.data.path.append("/tmp/nltk_data")
for _pkg in ["punkt", "stopwords", "vader_lexicon", "averaged_perceptron_tagger"]:
    try:
        nltk.download(_pkg, download_dir="/tmp/nltk_data", quiet=True)
    except Exception:
        pass
# Safely initialise NLTK tools — used in /check
_nltk_ready = False
try:
    stemmer    = PorterStemmer()
    stop_words = set(stopwords.words("english"))
    sia        = SentimentIntensityAnalyzer()
    _nltk_ready = True
except Exception as _e:
    print(f"⚠️ NLTK initialisation failed: {_e}")
    stemmer = stop_words = sia = None
# ROUTES
@app.route("/")
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("dashboard"))
    error = None
    success = request.args.get("success")  # FIX: read success message from register redirect
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()
        finally:
            conn.close()
        if user and user["password"] == hash_password(password):
            session["user"]      = username
            session["user_name"] = user["name"]
            session["user_role"] = user["role"]
            next_page = request.args.get("next")
            return redirect(next_page) if next_page else redirect(url_for("dashboard"))
        else:
            # Generic message avoids username enumeration
            error = "Invalid username or password"
    return render_template("login.html", error=error, success=success)
@app.route("/register", methods=["GET", "POST"])
def register():
    if "user" in session:
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        username         = request.form.get("username", "").strip()
        password         = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        name             = request.form.get("name", "").strip()
        if not all([username, password, confirm_password, name]):
            error = "All fields are required"
        elif password != confirm_password:
            error = "Passwords do not match"
        elif len(password) < 6:
            error = "Password must be at least 6 characters"
        else:
            conn = get_db()
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
                if cursor.fetchone():
                    error = "Username already exists"
                else:
                    cursor.execute(
                        "INSERT INTO users (username, password, name, role, joined) VALUES (?, ?, ?, ?, ?)",
                        (username, hash_password(password), name, "user", datetime.now().strftime("%Y-%m-%d"))
                    )
                    conn.commit()
                    # Email sent *after* conn is safely closed
            finally:
                conn.close()
            if not error:
                send_registration_email(username, name, "user")
                # FIX: redirect properly so browser URL changes to /login
                return redirect(url_for("login", success="Registration successful! Please login."))
    return render_template("register.html", error=error)
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))
@app.route("/dashboard")
@login_required
def dashboard():
    history = get_user_history()
    total   = len(history)
    fake    = sum(1 for h in history if h.get("is_fake"))
    genuine = total - fake
    stats = {
        "total":           total,
        "fake":            fake,
        "genuine":         genuine,
        "fake_percent":    round((fake    / total) * 100, 1) if total > 0 else 0,
        "genuine_percent": round((genuine / total) * 100, 1) if total > 0 else 0,
    }
    return render_template(
        "dashboard_advanced.html",
        user_name=session.get("user_name"),
        user_role=session.get("user_role"),
        history=history[:10],
        **stats,
    )
@app.route("/check", methods=["GET", "POST"])
@login_required
def check():
    result = None
    if request.method == "POST":
        review_text = request.form.get("review", "").strip()
        if not review_text:
            flash("Please enter a review to check.", "warning")
            return render_template("check.html",
                user_name=session.get("user_name"),
                user_role=session.get("user_role"),
                result=None)
        # ── Text metrics (always available) ──────────────────────────────
        word_count   = len(review_text.split())
        unique_words = round(len(set(review_text.lower().split())) / max(word_count, 1) * 100, 1)
        exclamations = review_text.count("!")
        caps_ratio   = round(sum(c.isupper() for c in review_text) / max(len(review_text), 1) * 100, 1)
        extreme_list = {
            "worst", "best", "terrible", "amazing", "horrible", "perfect",
            "awful", "fantastic", "garbage", "excellent", "pathetic", "outstanding"
        }
        extreme_words = sum(1 for w in review_text.lower().split() if w.strip(".,!?") in extreme_list)
        # ── PRIMARY: Claude AI analysis ───────────────────────────────────
        ai_verdict    = None
        ai_confidence = None
        ai_reason     = ""
        red_flags     = []
        green_flags   = []
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if api_key:
            try:
                import json as _json, urllib.request as _req, urllib.error as _uerr
                prompt = f"""You are an expert review fraud detection system. Analyze the following customer review and determine if it is FAKE or GENUINE.

Review to analyze:
\"\"\"{review_text}\"\"\"
Analyze based on:
- Language patterns (overly promotional, generic, repetitive, bot-like)
- Emotional authenticity (extreme praise/criticism without specifics)
- Detail level (vague vs specific personal experience)
- Writing style (natural vs templated)
- Contextual coherence
Respond ONLY with a valid JSON object in this exact format (no markdown, no extra text):
{{
  "verdict": "FAKE" or "GENUINE",
  "confidence": <integer 0-100>,
  "reason": "<one concise sentence explaining the verdict>",
  "red_flags": ["<specific flag>", "<specific flag>"],
  "green_flags": ["<specific indicator>", "<specific indicator>"]
}}"""
                payload = _json.dumps({
                    "model": "claude-sonnet-4-20250514",
                    "max_tokens": 400,
                    "messages": [{"role": "user", "content": prompt}]
                }).encode("utf-8")
                http_req = _req.Request(
                    "https://api.anthropic.com/v1/messages",
                    data=payload,
                    headers={
                        "Content-Type": "application/json",
                        "x-api-key": api_key,
                        "anthropic-version": "2023-06-01"
                    },
                    method="POST"
                )
                with _req.urlopen(http_req, timeout=20) as resp:
                    raw = _json.loads(resp.read().decode("utf-8"))
                content_text = ""
                for block in raw.get("content", []):
                    if block.get("type") == "text":
                        content_text += block.get("text", "")
                content_text = content_text.strip()
                if "```" in content_text:
                    content_text = content_text.split("```")[1]
                    if content_text.startswith("json"):
                        content_text = content_text[4:]
                content_text = content_text.strip().rstrip("`").strip()
                ai_resp      = _json.loads(content_text)
                ai_verdict   = str(ai_resp.get("verdict", "GENUINE")).upper()
                if ai_verdict not in ("FAKE", "GENUINE"):
                    ai_verdict = "GENUINE"
                ai_confidence = max(0, min(100, int(ai_resp.get("confidence", 75))))
                ai_reason     = str(ai_resp.get("reason", ""))
                red_flags     = ai_resp.get("red_flags", []) or []
                green_flags   = ai_resp.get("green_flags", []) or []
            except Exception as _e:
                print(f"AI analysis error in /check: {_e}")
                ai_verdict = None  # fall through to rule-based
        # ── FALLBACK: rule-based heuristic (when no API key or AI failed) ─
        if ai_verdict is None:
            # Build basic flags — lower thresholds to catch more fake signals
            if exclamations > 1:
                red_flags.append(f"Excessive exclamation marks ({exclamations})")
            if caps_ratio > 10:
                red_flags.append(f"High use of capital letters ({caps_ratio}%)")
            if extreme_words >= 1:
                red_flags.append(f"Extreme/emotional language detected ({extreme_words} word(s))")
            if word_count <= 8:
                red_flags.append("Very short review — lacks detail")
            if unique_words < 50:
                red_flags.append(f"Low word diversity ({unique_words}%) — may be templated")

            # Check for generic/promotional phrases
            generic_phrases = [
                "highly recommend", "five stars", "love it", "best ever",
                "amazing product", "great value", "definitely recommend",
                "must buy", "absolutely love", "perfect product", "exceeded expectations",
                "will buy again", "worth every penny", "life changing", "game changer"
            ]
            review_lower = review_text.lower()
            generic_hits = sum(1 for p in generic_phrases if p in review_lower)
            if generic_hits >= 1:
                red_flags.append(f"Contains generic promotional phrases ({generic_hits} found)")

            # Check repetition — fake reviews often repeat same words
            words_list = [w.strip(".,!?") for w in review_text.lower().split() if len(w) > 3]
            word_freq = Counter(words_list)
            repeated = [w for w, c in word_freq.items() if c >= 3]
            if repeated:
                red_flags.append(f"Repetitive word usage: {', '.join(repeated[:3])}")

            # Positive flags
            if word_count > 30:
                green_flags.append(f"Detailed review with {word_count} words")
            if unique_words > 70:
                green_flags.append(f"High word diversity ({unique_words}%) — natural writing")
            if exclamations == 0:
                green_flags.append("No excessive punctuation")
            if extreme_words == 0 and generic_hits == 0:
                green_flags.append("No extreme or promotional language")

            # Score heuristically — each flag contributes; threshold lowered to 15
            fake_score = (
                (exclamations > 1) * 20 +
                (caps_ratio > 10) * 15 +
                (extreme_words >= 1) * 15 +
                (word_count <= 8) * 20 +
                (unique_words < 50) * 15 +
                generic_hits * 20 +
                len(repeated) * 10
            )
            ai_verdict    = "FAKE" if fake_score >= 15 else "GENUINE"
            ai_confidence = min(95, 45 + fake_score) if ai_verdict == "FAKE" else min(88, 65 - fake_score // 2)
            ai_reason     = (
                f"Heuristic analysis detected {len(red_flags)} suspicious signal(s). "
                "Add ANTHROPIC_API_KEY for AI-powered detection."
            )
        # ── Also try local ML model as secondary signal if available ──────
        fake_prob    = 0.0
        genuine_prob = 100.0
        if model and vectorizer and _nltk_ready:
            try:
                text_clean = re.sub("[^a-zA-Z]", " ", review_text.lower())
                words_stem = [stemmer.stem(w) for w in text_clean.split() if w and w not in stop_words]
                processed  = " ".join(words_stem)
                vec        = vectorizer.transform([processed])
                probs      = model.predict_proba(vec)[0]
                fake_prob    = round(probs[1] * 100, 2)
                genuine_prob = round(probs[0] * 100, 2)
            except Exception:
                pass
        is_fake    = ai_verdict == "FAKE"
        confidence = float(ai_confidence)
        result = {
            "text":                review_text,
            "review":              review_text[:100] if len(review_text) > 100 else review_text,
            "review_full":         review_text,
            "is_fake":             is_fake,
            "confidence":          confidence,
            "fake_probability":    fake_prob if fake_prob else (confidence if is_fake else 100 - confidence),
            "genuine_probability": genuine_prob if genuine_prob else (100 - confidence if is_fake else confidence),
            "explanation": {
                "summary": (
                    ai_reason or
                    f"This review has been classified as {'FAKE' if is_fake else 'GENUINE'} "
                    f"with {confidence}% confidence."
                ),
                "red_flags":   red_flags,
                "green_flags": green_flags,
                "technical_details": {
                    "word_count":      word_count,
                    "unique_words":    unique_words,
                    "sentiment_score": round((confidence if is_fake else -confidence), 1),
                    "exclamations":    exclamations,
                    "caps_ratio":      caps_ratio,
                    "extreme_words":   extreme_words,
                },
            },
        }
        add_to_history(
            result["review"],
            result["review_full"],
            1 if result["is_fake"] else 0,
            result["confidence"],
            session.get("user"),
        )
    return render_template("check.html",
        user_name=session.get("user_name"),
        user_role=session.get("user_role"),
        result=result)
@app.route("/history")
@login_required
def history_view():
    user_role = session.get("user_role")
    history   = get_user_history() if user_role == "admin" else get_user_history(session.get("user"))
    return render_template("history.html",
        user_name=session.get("user_name"),
        user_role=user_role,
        history=history)
@app.route("/analytics")
@login_required
def analytics():
    start_date_str = request.args.get("start_date")
    end_date_str   = request.args.get("end_date")
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date   = datetime.strptime(end_date_str,   "%Y-%m-%d")
        except ValueError:
            start_date = datetime.now() - timedelta(days=30)
            end_date   = datetime.now()
    else:
        end_date   = datetime.now()
        start_date = end_date - timedelta(days=30)
    history  = get_user_history()
    filtered = []
    for item in history:
        try:
            item_date = datetime.strptime(item["timestamp"], "%Y-%m-%d %H:%M:%S")
            if start_date <= item_date <= end_date:
                filtered.append(item)
        except (ValueError, KeyError):
            continue
    total   = len(filtered)
    fake    = sum(1 for h in filtered if h.get("is_fake"))
    genuine = total - fake
    user_stats: dict = {}
    for item in filtered:
        user = item.get("checked_by", "Unknown")
        if user not in user_stats:
            user_stats[user] = {"total": 0, "fake": 0, "genuine": 0}
        user_stats[user]["total"] += 1
        if item.get("is_fake"):
            user_stats[user]["fake"] += 1
        else:
            user_stats[user]["genuine"] += 1
    return render_template("analytics.html",
        user_name=session.get("user_name"),
        user_role=session.get("user_role"),
        total=total,
        fake=fake,
        genuine=genuine,
        history=filtered,
        user_stats=user_stats,
        start_date=start_date.strftime("%Y-%m-%d"),
        end_date=end_date.strftime("%Y-%m-%d"))
@app.route("/users")
@admin_required
def user_management():
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, name, role, joined FROM users ORDER BY id DESC")
        users = [dict(row) for row in cursor.fetchall()]
        cursor.execute("SELECT checked_by, COUNT(*) as total FROM history GROUP BY checked_by")
        review_counts = {row["checked_by"]: row["total"] for row in cursor.fetchall()}
    finally:
        conn.close()
    for u in users:
        u["review_count"] = review_counts.get(u["username"], 0)
    return render_template("user_management.html",
        user_name=session.get("user_name"),
        user_role=session.get("user_role"),
        users=users,
        current_user=session.get("user"))
@app.route("/users/delete/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
        row = cursor.fetchone()
        if row and row["username"] == session.get("user"):
            flash("You cannot delete your own account.", "error")
            return redirect(url_for("user_management"))
        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for("user_management"))

@app.route("/users/role/<int:user_id>", methods=["POST"])
@admin_required
def change_role(user_id):
    new_role = request.form.get("role")
    if new_role not in ("admin", "user"):
        flash("Invalid role specified.", "error")
        return redirect(url_for("user_management"))
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for("user_management"))
@app.route("/user-dashboard")
@login_required
def user_dashboard():
    """Alias for regular users — redirects to the main dashboard."""
    return redirect(url_for("dashboard"))
# EXPORT ROUTES  (linked from dashboard nav)
@app.route("/export/csv")
@login_required
def export_csv():
    """Export review history as CSV"""
    import csv
    import io
    from flask import Response
    history = get_user_history()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Review", "Result", "Confidence (%)", "Checked By", "Timestamp"])
    for row in history:
        writer.writerow([
            row.get("id", ""),
            row.get("review_full") or row.get("review", ""),
            "Fake" if row.get("is_fake") else "Genuine",
            row.get("confidence", ""),
            row.get("checked_by", ""),
            row.get("timestamp", ""),
        ])
    output.seek(0)
    filename = f"fraud_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
@app.route("/export/excel")
@login_required
def export_excel():
    """Export review history as Excel (.xlsx)"""
    import io
    from flask import Response
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("openpyxl is not installed. Run: pip install openpyxl", "error")
        return redirect(url_for("dashboard"))
    history = get_user_history()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fraud Detection History"
    headers = ["ID", "Review", "Result", "Confidence (%)", "Checked By", "Timestamp"]
    header_fill = PatternFill(start_color="00CC7A", end_color="00CC7A", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(history, 2):
        ws.cell(row=r, column=1, value=row.get("id", ""))
        ws.cell(row=r, column=2, value=row.get("review_full") or row.get("review", ""))
        ws.cell(row=r, column=3, value="Fake" if row.get("is_fake") else "Genuine")
        ws.cell(row=r, column=4, value=row.get("confidence", ""))
        ws.cell(row=r, column=5, value=row.get("checked_by", ""))
        ws.cell(row=r, column=6, value=row.get("timestamp", ""))
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"fraud_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
@app.route("/export/pdf")
@login_required
def export_pdf():
    """Export review history as PDF"""
    import io
    from flask import Response
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
    except ImportError:
        flash("reportlab is not installed. Run: pip install reportlab", "error")
        return redirect(url_for("dashboard"))
    history = get_user_history()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("Review Fraud Detection — History Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))
    table_data = [["#", "Review (truncated)", "Result", "Confidence", "Checked By", "Timestamp"]]
    for row in history:
        review_text = str(row.get("review_full") or row.get("review", ""))
        table_data.append([
            str(row.get("id", "")),
            review_text[:60] + ("..." if len(review_text) > 60 else ""),
            "FAKE" if row.get("is_fake") else "GENUINE",
            f"{row.get('confidence', 0):.1f}%",
            str(row.get("checked_by", "")),
            str(row.get("timestamp", "")),
        ])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#00cc7a")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.black),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, 0),  10),
        ("FONTSIZE",       (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING",        (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    output.seek(0)
    filename = f"fraud_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        output.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
# FIX: was GET — must be POST to prevent CSRF / accidental clears via link
@app.route("/clear-history", methods=["POST"])
@admin_required
def clear_history():
    conn = get_db()
    try:
        conn.execute("DELETE FROM history")
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for("history_view"))
# LIVE ACTIVITY API  (used by dashboard)
@app.route("/api/live-activity")
@login_required
def live_activity():
    """Return recent activity for the dashboard live feed."""
    limit = min(int(request.args.get("limit", 10)), 50)
    history = get_user_history()          # newest first
    recent  = history[:limit]
    total   = len(history)
    fake    = sum(1 for h in history if h.get("is_fake"))
    genuine = total - fake
    return jsonify({
        "total": total,
        "activity": [
            {
                "review":     (h.get("review_full") or h.get("review") or "")[:100],
                "is_fake":    bool(h.get("is_fake")),
                "confidence": round(float(h.get("confidence") or 0), 1),
                "checked_by": h.get("checked_by", ""),
                "timestamp":  h.get("timestamp", ""),
            }
            for h in recent
        ],
        "stats": {
            "total":           total,
            "fake":            fake,
            "genuine":         genuine,
            "fake_percent":    round(fake    / total * 100, 1) if total else 0,
            "genuine_percent": round(genuine / total * 100, 1) if total else 0,
        },
    })
# CL - PRIMARY REVIEW ANALYSIS ENGINE
import urllib.request
import urllib.error
@app.route("/api/ai-check", methods=["POST"])
@login_required
def ai_check():
    """
    Use Claude AI as the PRIMARY fraud detection engine.
    This replaces the unreliable local ML model for accurate results.
    """
    data = request.get_json(silent=True) or {}
    review_text = (data.get("review") or "").strip()
    if not review_text:
        return jsonify({"error": "No review text provided"}), 400
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY not set"}), 503
    prompt = f"""You are an expert review fraud detection system. Analyze the following customer review and determine if it is FAKE or GENUINE.

Review to analyze:
\"\"\"{review_text}\"\"\"
Analyze based on these criteria:
- Language patterns (overly promotional, generic, repetitive)
- Emotional authenticity (extreme praise/criticism without specifics)
- Detail level (vague vs specific experience)
- Writing style (natural vs bot-like)
- Contextual relevance
Respond ONLY with a valid JSON object in this exact format (no markdown, no extra text):
{{
  "verdict": "FAKE" or "GENUINE",
  "confidence": <integer 0-100>,
  "reason": "<one concise sentence explaining the verdict>",
  "red_flags": ["<flag1>", "<flag2>"],
  "green_flags": ["<flag1>", "<flag2>"]
}}"""
    try:
        import json as _json
        payload = _json.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 400,
            "messages": [{"role": "user", "content": prompt}]
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = _json.loads(resp.read().decode("utf-8"))

        # Extract text from response
        content_text = ""
        for block in raw.get("content", []):
            if block.get("type") == "text":
                content_text += block.get("text", "")
        # Clean and parse JSON
        content_text = content_text.strip()
        if content_text.startswith("```"):
            content_text = content_text.split("```")[1]
            if content_text.startswith("json"):
                content_text = content_text[4:]
        content_text = content_text.strip().rstrip("```").strip()
        result = _json.loads(content_text)
        # Validate required fields
        verdict = str(result.get("verdict", "GENUINE")).upper()
        if verdict not in ("FAKE", "GENUINE"):
            verdict = "GENUINE"
        confidence = max(0, min(100, int(result.get("confidence", 75))))
        reason = str(result.get("reason", "Analysis complete."))
        red_flags = result.get("red_flags", [])
        green_flags = result.get("green_flags", [])
        # Ensure lists
        if not isinstance(red_flags, list):
            red_flags = []
        if not isinstance(green_flags, list):
            green_flags = []
        return jsonify({
            "verdict": verdict,
            "confidence": confidence,
            "reason": reason,
            "red_flags": red_flags,
            "green_flags": green_flags
        })
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        print(f"Anthropic API HTTP error {e.code}: {err_body}")
        return jsonify({"error": f"API error {e.code}: check your API key and quota"}), 502
    except Exception as e:
        print(f"AI check error: {e}")
        return jsonify({"error": "AI analysis failed. Please try again."}), 500
@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404
@app.errorhandler(500)
def server_error(e):
    return render_template("404.html"), 500  # reuse or add a 500.html
# ==========================================
# RUN APP
if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(debug=debug, host="127.0.0.1", port=port)
